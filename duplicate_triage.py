"""
Duplicate Triage System for NORE Q/A Training Data

TWO-PHASE workflow designed for human-in-the-loop deduplication:

  Phase 1 (--triage):   Analyze all Q/A pairs → Excel workbook for human review
  Phase 2 (--apply):    Read human decisions from workbook → clean JSONL output

DESIGN PRINCIPLES (from panel review 2026-02-11):
  - NEVER auto-delete. All removal decisions require human confirmation.
  - Every record appears in the output workbook, not just duplicates.
  - Direct pairwise similarity only — NO transitive clustering.
    (Union-Find chaining destroyed 85% of data in previous versions)
  - Answer-aware comparison distinguishes true duplicates from answer conflicts.
  - Safety-critical conflicts are flagged for priority review.

Usage:
    # Phase 1: Generate triage workbook
    python duplicate_triage.py --triage --dir qa_jsonl --out triage_workbook.xlsx

    # Phase 2: Apply decisions from reviewed workbook
    python duplicate_triage.py --apply --workbook triage_workbook.xlsx --clean-out qa_clean.jsonl

Dependencies:
    pip install sentence-transformers pandas openpyxl tqdm

Part of the NORE Q/A Generation Pipeline (post-processing stage)
"""

import json
import argparse
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any, Set, Tuple
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- CONFIGURATION ---
DEFAULT_DIR = "qa_jsonl"
MODEL_NAME = 'sentence-transformers/all-mpnet-base-v2'

# Question similarity threshold
DEFAULT_SIMILARITY_THRESHOLD = 0.85

# Answer comparison thresholds (freeform answers are 1-3 words)
ANSWER_SAME_THRESHOLD = 0.80      # above = "same answer"
ANSWER_CONFLICT_THRESHOLD = 0.50  # below = "different answer" → conflict

# Safety terms that flag a conflict as priority review
# (subset of verification_qa.py SAFETY_WATCHLIST)
SAFETY_PRIORITY_TERMS = {
    "methotrexate", "cisplatin", "carboplatin", "oxaliplatin", "5-fu",
    "fluorouracil", "capecitabine", "pembrolizumab", "nivolumab",
    "warfarin", "vitamin k", "folate", "folic acid", "potassium",
    "refeeding syndrome", "hyperkalemia", "hypokalemia",
    "neutropenia", "cachexia", "renal", "hepatic",
    "drug-nutrient", "interaction", "contraindicated",
    "grapefruit", "st john's wort",
}


# ============================================================================
# FILE I/O
# ============================================================================

def parse_jsonl(file_path: Path) -> List[Dict]:
    """Parse JSONL file (handles both standard and '---'-separated format)."""
    records = []
    try:
        content = file_path.read_text(encoding='utf-8')
        blocks = content.split('---') if '---' in content else content.splitlines()

        for block in blocks:
            block = block.strip()
            if not block:
                continue
            try:
                data = json.loads(block)
                if isinstance(data, dict) and 'question' in data and 'answer' in data:
                    data['_source_file'] = file_path.name
                    records.append(data)
            except json.JSONDecodeError:
                continue
    except Exception as e:
        logging.error(f"Error reading {file_path.name}: {e}")
    return records


def get_confidence(record: Dict) -> Optional[float]:
    """Extract verification confidence from nested or flat structure."""
    v = record.get('verification', {})
    if isinstance(v, dict):
        c = v.get('confidence')
        if c is not None:
            return float(c)
    c = record.get('verification_confidence')
    return float(c) if c is not None else None


def check_safety_relevance(question: str, answer: str) -> bool:
    """Check if a Q/A pair involves safety-critical content."""
    combined = (question + " " + answer).lower()
    return any(term in combined for term in SAFETY_PRIORITY_TERMS)


# ============================================================================
# CLASSIFICATION
# ============================================================================

def classify_pair(q_score: float, a_score: float) -> str:
    """Classify a pair based on question AND answer similarity."""
    if a_score >= ANSWER_SAME_THRESHOLD:
        return "true_duplicate"
    elif a_score < ANSWER_CONFLICT_THRESHOLD:
        return "answer_conflict"
    else:
        return "near_duplicate"


# ============================================================================
# PHASE 1: TRIAGE — Generate Excel workbook
# ============================================================================

def run_triage(
    input_dir: str,
    threshold: float = DEFAULT_SIMILARITY_THRESHOLD,
    output_xlsx: str = "triage_workbook.xlsx",
    verbose: bool = True
) -> None:
    """
    Phase 1: Analyze all Q/A pairs and generate an Excel triage workbook.

    The workbook has two sheets:
      - "All Questions": Every record with labels, group IDs, and a Decision column
      - "Duplicate Pairs": Detailed pair comparisons for reference

    NO records are deleted. The human fills in the Decision column.
    """
    from sentence_transformers import SentenceTransformer, util

    root_path = Path(input_dir)
    if not root_path.exists():
        logging.error(f"Directory not found: {input_dir}")
        return

    # --- 1. Load all records ---
    if verbose:
        print(f"Scanning {root_path} for *_qa.jsonl files...")

    all_records = []
    files = sorted(root_path.glob("*_qa.jsonl"))
    if not files:
        logging.error(f"No *_qa.jsonl files found in {input_dir}")
        return

    for p in tqdm(files, desc="Loading", disable=not verbose):
        all_records.extend(parse_jsonl(p))

    if not all_records:
        logging.error("No Q/A records found.")
        return

    n = len(all_records)
    if verbose:
        print(f"Loaded {n} freeform Q/A pairs from {len(files)} files")

    # --- 2. Embed questions and answers ---
    if verbose:
        print(f"\nLoading model: {MODEL_NAME}")

    model = SentenceTransformer(MODEL_NAME)
    questions = [r['question'] for r in all_records]
    answers = [r.get('answer', '') for r in all_records]

    if verbose:
        print("Encoding answers...")

    # Note: question embeddings are computed internally by paraphrase_mining below,
    # so we only pre-compute answer embeddings here (used for pairwise answer comparison).
    a_embeddings = model.encode(answers, show_progress_bar=verbose, batch_size=64)

    # --- 3. Find similar pairs (with top_k limit) ---
    if verbose:
        print(f"\nFinding question pairs with similarity > {threshold}...")

    # Use paraphrase_mining with top_k to avoid O(n²) explosion
    raw_pairs = util.paraphrase_mining(
        model, questions,
        show_progress_bar=verbose,
        batch_size=32,
        top_k=10
    )

    # --- 4. Build pair data with answer comparison ---
    pairs_data = []
    # Track which records are involved in any pair
    involved_in_pair: Dict[int, List[int]] = defaultdict(list)  # idx -> [pair_row_numbers]

    for score_tensor, idx1, idx2 in raw_pairs:
        q_score = float(score_tensor)
        if q_score < threshold:
            continue

        rec1 = all_records[idx1]
        rec2 = all_records[idx2]

        # Compare answers
        a_sim = float(util.cos_sim(a_embeddings[idx1], a_embeddings[idx2])[0][0])
        classification = classify_pair(q_score, a_sim)

        # Check safety relevance for conflicts
        is_safety = False
        if classification == "answer_conflict":
            is_safety = (check_safety_relevance(rec1['question'], rec1.get('answer', '')) or
                         check_safety_relevance(rec2['question'], rec2.get('answer', '')))

        pair_idx = len(pairs_data)
        involved_in_pair[idx1].append(pair_idx)
        involved_in_pair[idx2].append(pair_idx)

        c1 = get_confidence(rec1)
        c2 = get_confidence(rec2)

        pairs_data.append({
            "pair_id": pair_idx + 1,
            "q_similarity": round(q_score, 4),
            "a_similarity": round(a_sim, 4),
            "classification": classification,
            "safety_flag": "⚠ SAFETY" if is_safety else "",
            "q1_idx": idx1,
            "q1_id": rec1.get('qa_id', f'idx_{idx1}'),
            "q1_question": rec1['question'],
            "q1_answer": rec1.get('answer', ''),
            "q1_confidence": round(c1, 3) if c1 is not None else None,
            "q1_file": rec1['_source_file'],
            "q2_idx": idx2,
            "q2_id": rec2.get('qa_id', f'idx_{idx2}'),
            "q2_question": rec2['question'],
            "q2_answer": rec2.get('answer', ''),
            "q2_confidence": round(c2, 3) if c2 is not None else None,
            "q2_file": rec2['_source_file'],
        })

    if verbose:
        print(f"Found {len(pairs_data)} similar pairs above threshold {threshold}")

    # --- 5. Assign duplicate groups (DIRECT pairs only, no transitive chains) ---
    # Each pair defines a group. A record can be in multiple groups.
    # We assign each record its HIGHEST-similarity group membership.
    record_best_pair: Dict[int, Dict] = {}  # idx -> best pair info

    for p in pairs_data:
        for idx_key, other_key in [("q1_idx", "q2_idx"), ("q2_idx", "q1_idx")]:
            idx = p[idx_key]
            other_idx = p[other_key]
            if idx not in record_best_pair or p["q_similarity"] > record_best_pair[idx]["score"]:
                record_best_pair[idx] = {
                    "score": p["q_similarity"],
                    "pair_id": p["pair_id"],
                    "classification": p["classification"],
                    "other_idx": other_idx,
                    "a_similarity": p["a_similarity"],
                    "safety_flag": p["safety_flag"],
                }

    # --- 6. Build the "All Questions" sheet data ---
    all_questions_data = []
    for idx, rec in enumerate(all_records):
        conf = get_confidence(rec)
        pair_info = record_best_pair.get(idx)

        if pair_info:
            other_rec = all_records[pair_info["other_idx"]]
            status = pair_info["classification"]
            suggested = "REVIEW" if status == "answer_conflict" else "remove_candidate"

            # For true/near duplicates, suggest keeping the better one
            if status in ("true_duplicate", "near_duplicate"):
                my_conf = conf or 0.0
                other_conf = get_confidence(other_rec) or 0.0
                my_qlen = len(rec.get('question', ''))
                other_qlen = len(other_rec.get('question', ''))

                if (my_conf, my_qlen) >= (other_conf, other_qlen):
                    suggested = "keep (best in pair)"
                else:
                    suggested = "remove_candidate"
        else:
            status = "unique"
            suggested = "keep"

        all_questions_data.append({
            "row_num": idx + 1,
            "qa_id": rec.get('qa_id', f'idx_{idx}'),
            "question": rec['question'],
            "answer": rec.get('answer', ''),
            "source_file": rec['_source_file'],
            "chunk_id": rec.get('chunk_id', ''),
            "confidence": round(conf, 3) if conf is not None else None,
            "status": status,
            "suggested_action": suggested,
            "safety_flag": pair_info["safety_flag"] if pair_info else "",
            "q_similarity": pair_info["score"] if pair_info else None,
            "a_similarity": pair_info["a_similarity"] if pair_info else None,
            "most_similar_to": (all_records[pair_info["other_idx"]].get('qa_id', '')
                                if pair_info else ""),
            "pair_ref": pair_info["pair_id"] if pair_info else None,
            "human_decision": "",  # LEFT BLANK for human
        })

    # --- 7. Write Excel workbook ---
    if verbose:
        print(f"\nWriting triage workbook: {output_xlsx}")

    _write_triage_workbook(all_questions_data, pairs_data, output_xlsx, n, threshold)

    # --- 8. Print summary ---
    unique_count = sum(1 for r in all_questions_data if r["status"] == "unique")
    dup_count = sum(1 for r in all_questions_data if r["status"] == "true_duplicate")
    near_count = sum(1 for r in all_questions_data if r["status"] == "near_duplicate")
    conflict_count = sum(1 for r in all_questions_data if r["status"] == "answer_conflict")
    safety_count = sum(1 for r in all_questions_data if r["safety_flag"])

    print(f"\n{'='*60}")
    print(f"TRIAGE WORKBOOK GENERATED")
    print(f"{'='*60}")
    print(f"Total records:         {n}")
    print(f"  Unique (no matches): {unique_count}")
    print(f"  True duplicates:     {dup_count}")
    print(f"  Near duplicates:     {near_count}")
    print(f"  Answer conflicts:    {conflict_count}")
    if safety_count:
        print(f"  ⚠ Safety-flagged:    {safety_count}")
    print(f"\nSimilar pairs found:   {len(pairs_data)}")
    print(f"Threshold used:        {threshold}")
    print(f"\nWorkbook saved to: {output_xlsx}")
    print(f"\nNEXT STEPS:")
    print(f"  1. Open {output_xlsx} in Excel/LibreOffice")
    print(f"  2. Go to 'All Questions' sheet")
    print(f"  3. Sort by 'status' column to group duplicates and conflicts")
    print(f"  4. Fill in 'human_decision' column: keep / remove / merge")
    print(f"     - Start with ⚠ SAFETY flagged conflicts (most critical)")
    print(f"     - Then review answer_conflict rows")
    print(f"     - Then confirm true_duplicate removals")
    print(f"  5. Save the workbook")
    print(f"  6. Run: python duplicate_triage.py --apply --workbook {output_xlsx}")
    print(f"{'='*60}")


def _write_triage_workbook(
    all_questions: List[Dict],
    pairs: List[Dict],
    output_path: str,
    total_records: int,
    threshold: float
) -> None:
    """Write the Excel triage workbook with formatting and data validation."""

    wb = Workbook()

    # ---- Sheet 1: All Questions ----
    ws1 = wb.active
    ws1.title = "All Questions"

    # Column definitions
    q_columns = [
        ("Row #", 6),
        ("QA ID", 20),
        ("Question", 60),
        ("Answer", 20),
        ("Source File", 30),
        ("Chunk ID", 12),
        ("Confidence", 11),
        ("Status", 16),
        ("Suggested Action", 20),
        ("Safety Flag", 14),
        ("Q Similarity", 13),
        ("A Similarity", 13),
        ("Most Similar To", 20),
        ("Pair Ref #", 10),
        ("Human Decision", 16),
    ]

    # Header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, (col_name, col_width) in enumerate(q_columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Conditional fill colors
    fill_unique = PatternFill("solid", fgColor="E2EFDA")       # light green
    fill_true_dup = PatternFill("solid", fgColor="FCE4D6")     # light orange
    fill_near_dup = PatternFill("solid", fgColor="FFF2CC")     # light yellow
    fill_conflict = PatternFill("solid", fgColor="F8CECC")     # light red
    fill_safety = PatternFill("solid", fgColor="FF0000")       # red background
    fill_decision = PatternFill("solid", fgColor="D9E2F3")     # light blue (decision column)

    safety_font = Font(bold=True, color="FFFFFF")

    status_fills = {
        "unique": fill_unique,
        "true_duplicate": fill_true_dup,
        "near_duplicate": fill_near_dup,
        "answer_conflict": fill_conflict,
    }

    # Data rows
    field_order = [
        "row_num", "qa_id", "question", "answer", "source_file", "chunk_id",
        "confidence", "status", "suggested_action", "safety_flag",
        "q_similarity", "a_similarity", "most_similar_to", "pair_ref",
        "human_decision"
    ]

    for row_idx, record in enumerate(all_questions, 2):
        status = record.get("status", "")
        row_fill = status_fills.get(status, None)

        for col_idx, field in enumerate(field_order, 1):
            value = record.get(field, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "question"))

            # Apply status-based fill to the status column
            if field == "status" and row_fill:
                cell.fill = row_fill

            # Safety flag formatting
            if field == "safety_flag" and value:
                cell.fill = fill_safety
                cell.font = safety_font

            # Decision column highlight
            if field == "human_decision":
                cell.fill = fill_decision

            # Number formatting
            if field in ("confidence", "q_similarity", "a_similarity") and value is not None:
                cell.number_format = '0.000'

    # Add data validation dropdown for Human Decision column
    decision_col = len(q_columns)  # Last column
    dv = DataValidation(
        type="list",
        formula1='"keep,remove,merge,review_later"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Decision",
        error="Please choose: keep, remove, merge, or review_later"
    )
    dv.sqref = f"{get_column_letter(decision_col)}2:{get_column_letter(decision_col)}{len(all_questions)+1}"
    ws1.add_data_validation(dv)

    # Freeze header row
    ws1.freeze_panes = "A2"

    # Auto-filter
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(q_columns))}{len(all_questions)+1}"

    # ---- Sheet 2: Duplicate Pairs ----
    ws2 = wb.create_sheet("Duplicate Pairs")

    pair_columns = [
        ("Pair #", 8),
        ("Q Similarity", 13),
        ("A Similarity", 13),
        ("Classification", 16),
        ("Safety", 12),
        ("Q1 ID", 20),
        ("Q1 Question", 50),
        ("Q1 Answer", 20),
        ("Q1 Confidence", 13),
        ("Q1 File", 28),
        ("Q2 ID", 20),
        ("Q2 Question", 50),
        ("Q2 Answer", 20),
        ("Q2 Confidence", 13),
        ("Q2 File", 28),
    ]

    for col_idx, (col_name, col_width) in enumerate(pair_columns, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = col_width

    pair_field_order = [
        "pair_id", "q_similarity", "a_similarity", "classification", "safety_flag",
        "q1_id", "q1_question", "q1_answer", "q1_confidence", "q1_file",
        "q2_id", "q2_question", "q2_answer", "q2_confidence", "q2_file",
    ]

    for row_idx, pair in enumerate(pairs, 2):
        cls = pair.get("classification", "")
        row_fill = status_fills.get(cls, None)

        for col_idx, field in enumerate(pair_field_order, 1):
            value = pair.get(field, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(field in ("q1_question", "q2_question")))

            if field == "classification" and row_fill:
                cell.fill = row_fill
            if field == "safety_flag" and value:
                cell.fill = fill_safety
                cell.font = safety_font
            if field in ("q_similarity", "a_similarity", "q1_confidence", "q2_confidence") and value is not None:
                cell.number_format = '0.000'

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(pair_columns))}{len(pairs)+1}"

    # ---- Sheet 3: Instructions ----
    ws3 = wb.create_sheet("Instructions")
    ws3.column_dimensions['A'].width = 100

    instructions = [
        "NORE Q/A DUPLICATE TRIAGE WORKBOOK",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Total records: {total_records}  |  Similarity threshold: {threshold}",
        "",
        "═══════════════════════════════════════════════════════",
        "HOW TO USE THIS WORKBOOK",
        "═══════════════════════════════════════════════════════",
        "",
        "SHEET 1: 'All Questions'",
        "  Every Q/A record from your dataset with similarity labels.",
        "  The 'Human Decision' column (light blue) is where you mark your choices.",
        "",
        "  STATUS COLORS:",
        "    🟢 Green  = Unique (no similar matches found) → auto-keep",
        "    🟠 Orange = True Duplicate (same Q + same A) → safe to remove one",
        "    🟡 Yellow = Near Duplicate (same Q + close A) → likely remove, review first",
        "    🔴 Red    = Answer Conflict (same Q + DIFFERENT A) → must review",
        "    ⚠ Red BG  = Safety-critical conflict (drug-nutrient, toxicity) → review FIRST",
        "",
        "  DECISION OPTIONS (dropdown in Human Decision column):",
        "    keep         — include in final clean dataset",
        "    remove        — exclude from final clean dataset",
        "    merge         — flag for manual merge (combine best elements)",
        "    review_later  — defer decision",
        "",
        "  WORKFLOW:",
        "    1. Sort by 'Status' column to group similar records together",
        "    2. START with ⚠ SAFETY flagged conflicts (most critical)",
        "    3. Review answer_conflict rows — decide which answer is correct",
        "    4. Confirm true_duplicate removals (check suggested_action)",
        "    5. Mark unique records as 'keep' (or leave blank = keep by default)",
        "    6. Save the workbook",
        "",
        "SHEET 2: 'Duplicate Pairs'",
        "  Detailed side-by-side comparison of each similar pair.",
        "  Use the 'Pair Ref #' column in Sheet 1 to cross-reference.",
        "",
        "═══════════════════════════════════════════════════════",
        "APPLYING DECISIONS",
        "═══════════════════════════════════════════════════════",
        "",
        "After reviewing, run:",
        f"  python duplicate_triage.py --apply --workbook {Path(output_path).name} --dir {DEFAULT_DIR}",
        "",
        "This reads your decisions and produces a single clean JSONL file.",
        "Records with no decision or 'keep' are included.",
        "Records marked 'remove' are excluded.",
        "Records marked 'merge' or 'review_later' are reported for follow-up.",
        "",
        "═══════════════════════════════════════════════════════",
        "CLINICAL REVIEW GUIDELINES (from panel)",
        "═══════════════════════════════════════════════════════",
        "",
        "When reviewing ANSWER CONFLICTS:",
        "  1. Check both answers against authoritative guidelines (ESPEN, ASPEN, ASCO)",
        "  2. If both answers are valid for different contexts, keep BOTH with clarified questions",
        "  3. If one answer is clearly wrong, mark it 'remove' and the correct one 'keep'",
        "  4. For safety-critical conflicts: always verify against primary sources",
        "",
        "When reviewing NEAR DUPLICATES:",
        "  1. Keep the version with higher verification confidence",
        "  2. If confidence is equal, keep the more specific/detailed question",
        "  3. If questions are meaningfully different despite similarity, keep both",
    ]

    inst_font = Font(size=11, name="Arial")
    title_font = Font(size=14, bold=True, name="Arial")
    section_font = Font(size=11, bold=True, name="Arial")

    for row_idx, line in enumerate(instructions, 1):
        cell = ws3.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = title_font
        elif line.startswith("═"):
            cell.font = section_font
        elif line.startswith("  ") and "→" in line:
            cell.font = Font(size=11, name="Arial", italic=True)
        else:
            cell.font = inst_font

    # Save
    wb.save(output_path)


# ============================================================================
# PHASE 2: APPLY — Read decisions, produce clean JSONL
# ============================================================================

def apply_decisions(
    workbook_path: str,
    input_dir: str,
    output_file: str = "qa_clean.jsonl",
    verbose: bool = True
) -> None:
    """
    Phase 2: Read human decisions from triage workbook, produce clean JSONL.

    Rules:
      - 'keep' or blank → include in output
      - 'remove' → exclude
      - 'merge' → exclude but log for manual follow-up
      - 'review_later' → include but log warning
    """
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        logging.error(f"Workbook not found: {workbook_path}")
        return

    # Read decisions from workbook
    if verbose:
        print(f"Reading decisions from {workbook_path}...")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["All Questions"]

    # Build set of (qa_id, source_file) tuples to remove
    # Using tuples because qa_id alone is NOT unique across PDFs
    # (e.g. "ff-0-0" can appear in paper1_qa.jsonl AND paper2_qa.jsonl)
    remove_keys: Set[Tuple[str, str]] = set()
    merge_keys: Set[Tuple[str, str]] = set()
    review_later_keys: Set[Tuple[str, str]] = set()
    keep_count = 0
    no_decision_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 15:
            continue
        qa_id = str(row[1]) if row[1] else None      # Column B = QA ID
        source_file = str(row[4]) if row[4] else ""   # Column E = Source File
        decision = str(row[14]).strip().lower() if row[14] else ""  # Column O = Human Decision

        if not qa_id:
            continue

        key = (qa_id, source_file)

        if decision == "remove":
            remove_keys.add(key)
        elif decision == "merge":
            merge_keys.add(key)
        elif decision == "review_later":
            review_later_keys.add(key)
            keep_count += 1  # Include but warn
        elif decision in ("keep", ""):
            keep_count += 1
            if decision == "":
                no_decision_count += 1

    wb.close()

    if verbose:
        print(f"Decisions parsed:")
        print(f"  Keep (explicit + no decision): {keep_count}")
        print(f"  Remove: {len(remove_keys)}")
        print(f"  Merge (logged for follow-up): {len(merge_keys)}")
        print(f"  Review later (included with warning): {len(review_later_keys)}")
        if no_decision_count > 0:
            print(f"  ({no_decision_count} records had no decision → defaulting to keep)")

    # Load all records from source JSONL files
    root_path = Path(input_dir)
    all_records = []
    files = sorted(root_path.glob("*_qa.jsonl"))
    for p in files:
        all_records.extend(parse_jsonl(p))

    # Filter and write clean output
    out_path = Path(output_file)
    kept = 0
    removed = 0

    with open(out_path, 'w', encoding='utf-8') as f:
        for rec in all_records:
            qa_id = rec.get('qa_id', '')
            source_file = rec.get('_source_file', '')
            key = (qa_id, source_file)

            if key in remove_keys or key in merge_keys:
                removed += 1
                continue

            # Strip internal tracking fields
            clean = {k: v for k, v in rec.items() if not k.startswith('_')}
            f.write(json.dumps(clean, ensure_ascii=False) + "\n")
            kept += 1

    # Write merge log if needed
    if merge_keys:
        merge_log = out_path.parent / "merge_candidates.txt"
        with open(merge_log, 'w') as f:
            f.write("Records marked for merge — manual action required:\n\n")
            for qa_id, src_file in sorted(merge_keys):
                matching = [r for r in all_records
                            if r.get('qa_id') == qa_id and r.get('_source_file') == src_file]
                for m in matching:
                    f.write(f"  ID: {qa_id}\n")
                    f.write(f"  Q:  {m['question']}\n")
                    f.write(f"  A:  {m.get('answer', '')}\n")
                    f.write(f"  File: {m.get('_source_file', '')}\n\n")
        if verbose:
            print(f"\nMerge candidates logged to: {merge_log}")

    if verbose:
        print(f"\n{'='*60}")
        print(f"CLEAN DATASET GENERATED")
        print(f"{'='*60}")
        print(f"Input:    {len(all_records)} records")
        print(f"Removed:  {removed}")
        print(f"Output:   {kept} records → {out_path}")
        if review_later_keys:
            print(f"  {len(review_later_keys)} 'review_later' records included (resolve in next pass)")
        print(f"{'='*60}")


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Two-phase duplicate triage: Excel workbook for human review → clean JSONL.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Phase 1 — Generate triage workbook:
  python duplicate_triage.py --triage --dir qa_jsonl --out triage_workbook.xlsx

Phase 2 — Apply human decisions:
  python duplicate_triage.py --apply --workbook triage_workbook.xlsx --dir qa_jsonl

Part of the NORE Q/A Generation Pipeline (post-processing stage).
        """
    )

    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--triage", action="store_true",
                       help="Phase 1: Generate triage workbook for human review")
    mode.add_argument("--apply", action="store_true",
                       help="Phase 2: Apply decisions from reviewed workbook")
    mode.add_argument("--stats-only", action="store_true",
                       help="Show dataset statistics only")

    parser.add_argument("--dir", default=DEFAULT_DIR,
                        help=f"Directory containing *_qa.jsonl files (default: {DEFAULT_DIR})")
    parser.add_argument("--threshold", type=float, default=DEFAULT_SIMILARITY_THRESHOLD,
                        help=f"Question similarity threshold (default: {DEFAULT_SIMILARITY_THRESHOLD})")
    parser.add_argument("--out", default="triage_workbook.xlsx",
                        help="Output workbook filename (default: triage_workbook.xlsx)")
    parser.add_argument("--workbook", default="triage_workbook.xlsx",
                        help="Path to reviewed workbook (for --apply)")
    parser.add_argument("--clean-out", default="qa_clean.jsonl",
                        help="Output clean JSONL filename (default: qa_clean.jsonl)")
    parser.add_argument("--quiet", action="store_true",
                        help="Suppress progress output")

    args = parser.parse_args()

    if args.triage:
        run_triage(args.dir, args.threshold, args.out, verbose=not args.quiet)
    elif args.apply:
        apply_decisions(args.workbook, args.dir, args.clean_out, verbose=not args.quiet)
    elif args.stats_only:
        # Reuse simple stats from original
        root_path = Path(args.dir)
        all_records = []
        for p in sorted(root_path.glob("*_qa.jsonl")):
            all_records.extend(parse_jsonl(p))
        print(f"Total records: {len(all_records)}")
        files = set(r['_source_file'] for r in all_records)
        for fn in sorted(files):
            count = sum(1 for r in all_records if r['_source_file'] == fn)
            print(f"  {fn}: {count}")
